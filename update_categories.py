"""
legal_forms 테이블에 cat_large / cat_medium / cat_small / tag1 / tag2 / tag3 값을 채웁니다.
엑셀(outputs/법률서식 매핑, 태그 추가.xlsx) 1시트의 서식명을 기준으로 매핑합니다.

매핑 전략 (순서대로 적용):
  1. 정확 매핑
  2. 엑셀 타이틀의 [xxx] 접두사 제거 후 매핑
  3. 숫자 접두사(예: 1142) 기준 퍼지 매핑
  4. 브라켓 제거 후 퍼지 매핑 (cutoff 0.80)
  5. 특수 케이스 (OCR 오타, 공백 차이 등)

실행 전 migration_add_categories.sql 을 Supabase SQL Editor에서 먼저 실행하세요.
"""

import sys, os, re, json, requests
from difflib import SequenceMatcher, get_close_matches

sys.stdout.reconfigure(encoding="utf-8")

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

import openpyxl

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY") or os.environ["SUPABASE_KEY"]
EXCEL_PATH   = os.path.join(os.path.dirname(__file__), "outputs", "법률서식 매핑, 태그 추가.xlsx")

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal",
}

CHUNK = 80   # IN 절 최대 ID 수


# ── 헬퍼 ─────────────────────────────────────────────────────────────────────

def strip_bracket_prefix(t: str) -> str:
    """'[가사] 제목' → '제목'"""
    return re.sub(r"^\[[^\]]+\]\s*", "", t).strip()


def similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()


# ── 1. 엑셀 로드 ─────────────────────────────────────────────────────────────

print("엑셀 로드 중…")
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.worksheets[0]

# 헤더: 파일ID, 수집처, 원본_대분류, 서식명, 통합_대분류, 중분류, 소분류, 태그1, 태그2, 태그3, …
excel_data: dict[str, dict] = {}
for r in range(2, ws.max_row + 1):
    title = ws.cell(r, 4).value
    if not title:
        continue
    title = title.strip()
    excel_data[title] = {
        "cat_large":  ws.cell(r, 5).value,
        "cat_medium": ws.cell(r, 6).value,
        "cat_small":  ws.cell(r, 7).value,
        "tag1":       ws.cell(r, 8).value,
        "tag2":       ws.cell(r, 9).value,
        "tag3":       ws.cell(r, 10).value,
    }

print(f"  엑셀 서식 수: {len(excel_data):,}개")

excel_titles = set(excel_data.keys())


# ── 2. DB에서 id + title 전체 로드 ───────────────────────────────────────────

print("DB legal_forms 로드 중…")
all_rows: list[dict] = []
offset = 0
while True:
    r = requests.get(
        f"{SUPABASE_URL}/rest/v1/legal_forms?select=id,title&limit=1000&offset={offset}",
        headers=HEADERS,
    )
    r.raise_for_status()
    batch = r.json()
    if not batch:
        break
    all_rows.extend(batch)
    offset += len(batch)
    if len(batch) < 1000:
        break

print(f"  DB 행 수: {len(all_rows):,}개")
db_titles = {row["title"] for row in all_rows}


# ── 3. 매핑 구성 (db_title → excel_title) ────────────────────────────────────

print("매핑 구성 중…")
mapping: dict[str, str] = {}   # db_title → excel_title

# 전략 1: 정확
for t in excel_titles & db_titles:
    mapping[t] = t

# 전략 2: 브라켓 제거
for excel_t in excel_titles - set(mapping.values()):
    stripped = strip_bracket_prefix(excel_t)
    if stripped in db_titles and stripped not in mapping:
        mapping[stripped] = excel_t

# 전략 3: 숫자 접두사 + 유사도
mapped_db  = set(mapping.keys())
remaining_db = db_titles - mapped_db
remaining_excel = excel_titles - set(mapping.values())

for excel_t in list(remaining_excel):
    m = re.match(r"^(\d+[\-\d]*)\s+", excel_t)
    if not m:
        continue
    num_prefix = m.group(1)
    best_db, best_score = None, 0.0
    for db_t in remaining_db:
        if db_t.startswith(num_prefix + " ") or db_t.startswith(num_prefix + "-"):
            s = similarity(excel_t, db_t)
            if s > best_score:
                best_score, best_db = s, db_t
    if best_db and best_score > 0.75:
        mapping[best_db] = excel_t
        remaining_db.discard(best_db)

# 전략 4: 브라켓 제거 후 퍼지 (cutoff 0.80)
mapped_db   = set(mapping.keys())
remaining_db2 = list(db_titles - mapped_db)
remaining_excel2 = excel_titles - set(mapping.values())

for excel_t in list(remaining_excel2):
    stripped = strip_bracket_prefix(excel_t)
    matches = get_close_matches(stripped, remaining_db2, n=1, cutoff=0.80)
    if matches:
        mapping[matches[0]] = excel_t
        remaining_db2.remove(matches[0])

# 전략 5: 특수 케이스 (OCR 오타 등)
SPECIAL = {
    # db_title                                         → excel_title
    "휴대전화를 통한 정보수신 신청서":
        "[민사] 휴대전화들 동한 정보수신 신청서",
    "재산목록(강제집행 - 재산명시, A3211)":
        "[신청] 재산목록(강제집행-재산명시, A3211)",
}
for db_t, excel_t in SPECIAL.items():
    if db_t in db_titles and excel_t in excel_data and db_t not in mapping:
        mapping[db_t] = excel_t

print(f"  매핑 완료: {len(mapping):,}개 DB 타이틀")
print(f"  미매핑 DB 타이틀: {len(db_titles) - len(mapping):,}개")


# ── 4. id → 분류 데이터 매핑 ─────────────────────────────────────────────────

id_to_payload: dict[int, dict] = {}
for row in all_rows:
    db_t = row["title"]
    if db_t not in mapping:
        continue
    excel_t = mapping[db_t]
    data    = excel_data[excel_t]
    # None 값은 보내지 않아도 되지만 명시적으로 포함
    id_to_payload[row["id"]] = data


# ── 5. 분류값별로 그룹화 → 청크 단위 PATCH ────────────────────────────────────

from collections import defaultdict

payload_groups: dict[str, list[int]] = defaultdict(list)
for row_id, payload in id_to_payload.items():
    key = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    payload_groups[key].append(row_id)

print(f"\n총 {len(id_to_payload):,}개 행 업데이트 예정 ({len(payload_groups):,}개 고유 분류값)")
print("업데이트 시작…")

total_updated = 0
total_failed  = 0

for key_str, ids in payload_groups.items():
    payload = json.loads(key_str)
    # 청크로 나눠서 PATCH
    for i in range(0, len(ids), CHUNK):
        chunk = ids[i : i + CHUNK]
        id_filter = f"id=in.({','.join(str(x) for x in chunk)})"
        resp = requests.patch(
            f"{SUPABASE_URL}/rest/v1/legal_forms?{id_filter}",
            headers=HEADERS,
            json=payload,
        )
        if resp.status_code in (200, 204):
            total_updated += len(chunk)
        else:
            total_failed += len(chunk)
            print(f"  [실패] ids={chunk[:3]}… status={resp.status_code} body={resp.text[:200]}")

print(f"\n완료: 성공 {total_updated:,}건 / 실패 {total_failed:,}건")
