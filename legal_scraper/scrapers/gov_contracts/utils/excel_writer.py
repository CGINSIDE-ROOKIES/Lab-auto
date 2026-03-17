"""
Excel 결과 저장 유틸리티
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from ..base_scraper import FormItem

COLUMNS = [
    "일련번호",
    "부처명",
    "서식제목",
    "첨부파일명",
    "파일확장자",
    "담당부서",
    "등록일",
    "출처URL",
    "다운로드URL",
    "로컬경로",
    "수집일시",
]

# 열별 최소/최대 너비 (글자 단위)
_COL_MIN_WIDTH = 6
_COL_MAX_WIDTH = 60

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _item_to_row(seq: int, item: FormItem) -> list:
    return [
        seq,
        item.ministry,
        item.title,
        item.file_name,
        item.file_ext,
        item.department,
        item.registered_date,
        item.source_url,
        item.file_url,
        item.local_path,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ]


def _autofit_columns(ws) -> None:
    """모든 열의 너비를 내용 기준으로 자동 조정한다."""
    col_widths: dict[int, float] = {}

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            # 한글은 영문 대비 약 1.8배 너비 차지
            text = str(cell.value)
            width = sum(1.8 if ord(c) > 127 else 1.0 for c in text)
            col_widths[cell.column] = max(col_widths.get(cell.column, 0), width)

    for col_idx, width in col_widths.items():
        adjusted = min(max(width + 2, _COL_MIN_WIDTH), _COL_MAX_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


def save_to_excel(items: list[FormItem], output_path: str | Path) -> None:
    """
    items를 Excel 파일로 저장한다.
    - 파일이 없으면 생성, 있으면 데이터를 append한다.
    - 첫 열은 일련번호 (append 시 기존 마지막 번호에서 이어서 증가)
    - 헤더 행: 파란 배경(4472C4) + 흰 글씨 bold
    - 열 너비: 내용 기준 자동 조정
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if output_path.exists():
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        # 기존 마지막 일련번호 파악 (헤더 제외)
        last_seq = ws.max_row - 1  # 헤더 1행 빼기
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "gov_contracts"

        # 헤더 행 작성
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _HEADER_ALIGN

        last_seq = 0

    # 데이터 append
    for i, item in enumerate(items, start=last_seq + 1):
        ws.append(_item_to_row(i, item))

    _autofit_columns(ws)
    wb.save(output_path)
