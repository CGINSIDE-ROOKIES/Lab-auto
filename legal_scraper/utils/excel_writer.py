"""
엑셀 저장 유틸리티.
- write_raw()        : 수집 결과 저장 (단일 시트)
- write_classified() : 분류 결과 저장 (시트1: 전체 / 시트2: 검토필요)
- to_bytes_raw()       / to_bytes_classified() : Streamlit 다운로드용 BytesIO 반환
"""
import io
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import RAW_COLUMNS, CLASSIFIED_COLUMNS

# 배경색
_BLUE   = "4472C4"
_GREEN  = "C6EFCE"
_YELLOW = "FFEB9C"
_RED    = "FFC7CE"


def _header_style(ws):
    fill = PatternFill("solid", fgColor=_BLUE)
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws):
    for col_idx, col_cells in enumerate(ws.columns, 1):
        w = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(int(w * 1.3) + 2, 80)


def _fill_row(ws_row, status: str):
    color = {"분류완료": _GREEN, "부분일치": _YELLOW, "검토필요": _RED}.get(status)
    if color:
        fill = PatternFill("solid", fgColor=color)
        for cell in ws_row:
            cell.fill = fill


# ── 수집 결과 ──────────────────────────────────────────────────────
def _build_raw_wb(rows: list[dict]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "법률서식"
    ws.append(RAW_COLUMNS)
    _header_style(ws)
    for i, r in enumerate(rows, 1):
        ws.append([i, r.get("수집처",""), r.get("대분류",""), r.get("중분류",""),
                   r.get("서식제목",""), r.get("파일형식",""), r.get("다운로드URL",""), r.get("수집일시","")])
    ws.freeze_panes = "A2"
    _auto_width(ws)
    return wb


def write_raw(rows: list[dict], path: Path) -> Path:
    wb = _build_raw_wb(rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def to_bytes_raw(rows: list[dict]) -> bytes:
    buf = io.BytesIO()
    _build_raw_wb(rows).save(buf)
    return buf.getvalue()


# ── 분류 결과 ──────────────────────────────────────────────────────
def _build_classified_wb(rows: list[dict]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    # 시트1: 전체
    ws1 = wb.active
    ws1.title = "전체 분류 결과"
    ws1.append(CLASSIFIED_COLUMNS)
    _header_style(ws1)
    for i, r in enumerate(rows, 1):
        ws1.append([i,
            r.get("수집처",""), r.get("원본대분류",""), r.get("원본중분류",""),
            r.get("서식제목",""), r.get("파일형식",""),
            r.get("최종대분류",""), r.get("최종중분류",""), r.get("소분류",""), r.get("세분류",""),
            r.get("다운로드URL",""), r.get("수집일시",""), r.get("분류방법",""), r.get("분류상태",""),
        ])
        _fill_row(ws1[ws1.max_row], r.get("분류상태",""))
    ws1.freeze_panes = "A2"
    _auto_width(ws1)

    # 시트2: 검토필요
    ws2 = wb.create_sheet("검토필요")
    ws2.append(CLASSIFIED_COLUMNS)
    _header_style(ws2)
    review = [r for r in rows if r.get("분류상태") == "검토필요"]
    for i, r in enumerate(review, 1):
        ws2.append([i,
            r.get("수집처",""), r.get("원본대분류",""), r.get("원본중분류",""),
            r.get("서식제목",""), r.get("파일형식",""),
            r.get("최종대분류",""), r.get("최종중분류",""), r.get("소분류",""), r.get("세분류",""),
            r.get("다운로드URL",""), r.get("수집일시",""), r.get("분류방법",""), r.get("분류상태",""),
        ])
        _fill_row(ws2[ws2.max_row], "검토필요")
    ws2.freeze_panes = "A2"
    _auto_width(ws2)

    return wb


def write_classified(rows: list[dict], path: Path) -> Path:
    wb = _build_classified_wb(rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def to_bytes_classified(rows: list[dict]) -> bytes:
    buf = io.BytesIO()
    _build_classified_wb(rows).save(buf)
    return buf.getvalue()


# ── 데이터 병합 ────────────────────────────────────────────────────
import pandas as pd  # noqa: E402 (local import OK here)

EXPECTED_COLUMNS = ["일련번호", "수집처", "대분류", "중분류", "서식제목", "파일형식", "다운로드URL", "수집일시"]
DEDUP_KEYS = ["수집처", "대분류", "서식제목", "파일형식", "다운로드URL"]


def merge_files(base_df: pd.DataFrame, incr_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    기존 파일 DataFrame과 증분 파일 DataFrame을 병합한다.

    Returns
    -------
    merged_df : 병합 완료된 전체 DataFrame (일련번호 재부여)
    duplicates_df : 중복 행만 담은 DataFrame (없으면 빈 DataFrame)
    """
    # 중복 검사: 기존+증분 사이의 중복 (DEDUP_KEYS 5개 모두 일치)
    base_keys = base_df[DEDUP_KEYS].apply(tuple, axis=1)
    incr_keys = incr_df[DEDUP_KEYS].apply(tuple, axis=1)
    base_key_set = set(base_keys)
    dup_mask = incr_keys.isin(base_key_set)
    duplicates_df = incr_df[dup_mask].reset_index(drop=True)

    # 일련번호 재부여: 기존 최대 + 1 부터 순차
    max_seq = int(base_df["일련번호"].max()) if not base_df.empty else 0
    incr_reset = incr_df.copy()
    incr_reset["일련번호"] = range(max_seq + 1, max_seq + 1 + len(incr_reset))

    merged_df = pd.concat([base_df, incr_reset], ignore_index=True)
    return merged_df, duplicates_df


def merged_df_to_bytes(merged_df: pd.DataFrame) -> bytes:
    """병합된 DataFrame을 엑셀 bytes로 변환 (다운로드용)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "법률서식"
    ws.append(EXPECTED_COLUMNS)
    _header_style(ws)
    for _, row in merged_df.iterrows():
        ws.append([row.get(col, "") for col in EXPECTED_COLUMNS])
    ws.freeze_panes = "A2"
    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_merged(merged_df: pd.DataFrame, path: Path) -> Path:
    """병합된 DataFrame을 엑셀 파일로 저장."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "법률서식"
    ws.append(EXPECTED_COLUMNS)
    _header_style(ws)
    for _, row in merged_df.iterrows():
        ws.append([row.get(col, "") for col in EXPECTED_COLUMNS])
    ws.freeze_panes = "A2"
    _auto_width(ws)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
